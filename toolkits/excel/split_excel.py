from __future__ import annotations

from logging import Logger, getLogger
from pathlib import Path
from typing import Optional, Dict, List, Any

from openpyxl import load_workbook
from openpyxl.styles import Border, Side


class ExcelSplitter:
    """
    将单个 Excel 按首列(A列)的分组拆分为多个工作表，并在 H 列(第8列)汇总金额。

    行为与 `.data/split_excel.py` 一致：
    - 从第2行开始按 A 列的值分组；
    - 新表复制原表表头与样式；
    - 新表第一列为重新编号(1,2,3,... )；
    - 对 H 列(第8列)数值求和，合计写在数据最后一行+3 的 H 列单元格；
    - 保证数据区最后一行到 H 列有下边框；
    - 删除原始 sheet 仅保留新建的分组 sheet；
    - 输出文件名默认为原文件名加 `_split` 后缀。
    """

    def __init__(self, logger: Optional[Logger] = None):
        self.logger = logger or getLogger(__name__)

    def split_by_first_column(
        self,
        input_file: Path | str,
        output_file: Optional[Path | str] = None,
        amount_col: int = 8,
    ) -> Path:
        input_path = Path(input_file)
        if output_file is None:
            output_path = input_path.with_name(f"{input_path.stem}_split{input_path.suffix}")
        else:
            output_path = Path(output_file)

        if not input_path.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_path}")

        wb = load_workbook(str(input_path))
        ws = wb.active

        if ws.max_column < amount_col:
            raise ValueError(
                f"工作表至少需要 {amount_col} 列（H列），当前只有 {ws.max_column} 列"
            )

        groups: Dict[Any, List[int]] = {}
        for row in range(2, ws.max_row + 1):
            key = ws.cell(row=row, column=1).value
            if key is None:
                continue
            groups.setdefault(key, []).append(row)

        for key, rows in groups.items():
            # 生成 sheet 名称（安全处理、最长31字符）
            sheet_name = str(key)[:31]
            for char in "\\/?*[]:":
                sheet_name = sheet_name.replace(char, "_")
            new_ws = wb.create_sheet(title=sheet_name)

            # 列宽（与脚本保持一致，B=25, E=18）
            new_ws.column_dimensions["B"].width = 25
            new_ws.column_dimensions["E"].width = 18

            max_col = ws.max_column

            # 复制表头
            for col in range(1, max_col + 1):
                src = ws.cell(1, col)
                dst = new_ws.cell(1, col, value=src.value)
                if src.has_style:
                    dst.font = src.font.copy()
                    dst.border = src.border.copy()
                    dst.fill = src.fill.copy()
                    dst.number_format = src.number_format
                    dst.protection = src.protection.copy()
                    dst.alignment = src.alignment.copy()

            total = 0
            # 写入数据行，第一列为重新编号
            for i, old_row in enumerate(rows, start=2):
                for col in range(1, max_col + 1):
                    src = ws.cell(old_row, col)
                    if col == 1:
                        val = i - 1
                    else:
                        val = src.value
                        if col == amount_col and isinstance(val, (int, float)):
                            total += val
                    dst = new_ws.cell(i, col, value=val)
                    if src.has_style:
                        dst.font = src.font.copy()
                        dst.border = src.border.copy()
                        dst.fill = src.fill.copy()
                        dst.number_format = src.number_format
                        dst.protection = src.protection.copy()
                        dst.alignment = src.alignment.copy()

            # 确保数据最后一行到 H 列有下边框
            last_data_row = len(rows) + 1
            for col in range(1, amount_col + 1):
                cell = new_ws.cell(row=last_data_row, column=col)
                if cell.border:
                    cell.border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=Side(style="thin"),
                    )
                else:
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

            # 合计写在 len(rows) + 4 行的 H 列
            total_row = len(rows) + 4
            total_cell = new_ws.cell(row=total_row, column=amount_col, value=total)
            ref_cell = ws.cell(row=2, column=amount_col)
            if ref_cell.has_style:
                total_cell.number_format = ref_cell.number_format

        # 删除原始 sheet
        original_sheet_name = ws.title
        if original_sheet_name in wb.sheetnames:
            wb.remove(wb[original_sheet_name])

        # 保存
        wb.save(str(output_path))
        self.logger.info(f"处理完成，输出文件: {output_path}")
        return output_path


