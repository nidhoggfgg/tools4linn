from logging import Logger
from pathlib import Path
from openpyxl import load_workbook, Workbook
from typing import Optional

from toolkits.excel._utils import copy_sheet, find_all_excel_files
from toolkits.utils.naming import INDEXED_STRATEGY, NamingStrategy
from toolkits.utils.file_filter import FileFilterStrategy


class ExcelMerger:
    def __init__(
        self,
        logger: Logger,
        input_dir: Path,
        output_file: Path,
        sheet_naming_strategy: Optional[NamingStrategy] = None,
        file_filter_strategy: Optional[FileFilterStrategy] = None,
    ):
        self.logger = logger
        self.success_count = 0
        self.input_dir = input_dir
        self.output_file = output_file

        # Set up sheet naming strategy
        if sheet_naming_strategy is None:
            self.sheet_naming_strategy = INDEXED_STRATEGY
        else:
            self.sheet_naming_strategy = sheet_naming_strategy

        self.file_filter_strategy = file_filter_strategy

    def merge_excel(self):
        wb_out = Workbook()
        if wb_out.active is not None:
            wb_out.remove(wb_out.active)

        excel_files = find_all_excel_files(self.input_dir, self.file_filter_strategy)
        used_names = set()  # Track used sheet names to avoid duplicates

        for excel_file in excel_files:
            wb_in = load_workbook(excel_file, data_only=True)
            if len(wb_in.worksheets) == 0:
                self.logger.warning(f"excel file {excel_file} has no sheets")
                continue

            sheet_name = self.sheet_naming_strategy.generate_name(excel_file)
            if sheet_name in used_names:
                self.logger.error(f"sheet name {sheet_name} already exists")
                continue
            used_names.add(sheet_name)

            src_sheet = wb_in.worksheets[0]
            target_sheet = wb_out.create_sheet(title=sheet_name)
            copy_sheet(src_sheet, target_sheet)
