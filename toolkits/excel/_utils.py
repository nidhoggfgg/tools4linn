from pathlib import Path
from openpyxl.utils import get_column_letter
from typing import Optional

from toolkits.utils.file_filter import FileFilterStrategy


def copy_sheet(source_sheet, target_sheet):
    for row in source_sheet.iter_rows():
        src_row_dim = source_sheet.row_dimensions[row[0].row]
        if src_row_dim.height is not None:
            target_sheet.row_dimensions[row[0].row].height = src_row_dim.height

    for col in source_sheet.iter_cols():
        col_letter = get_column_letter(col[0].column)
        src_col_dim = source_sheet.column_dimensions[col_letter]
        if src_col_dim.width is not None:
            target_sheet.column_dimensions[col_letter].width = src_col_dim.width

    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()

    # merged_cells
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))


def find_all_excel_files(
    input_dir: Path, filter_strategy: Optional[FileFilterStrategy] = None
) -> list[Path]:
    """
    Find all .xlsx and .xls files in the input directory and its subdirectories,
    optionally filtered by a filtering strategy.

    Args:
        input_dir: Directory to search for Excel files
        filter_strategy: Optional filtering strategy to apply to files

    Returns:
        List of Excel file paths that pass the filter
    """
    all_excel_files = list(input_dir.rglob("*.xlsx")) + list(input_dir.rglob("*.xls"))

    filtered_files = []
    for file_path in all_excel_files:
        if filter_strategy is None:
            filtered_files.append(file_path)
            continue

        if filter_strategy.should_include(file_path):
            filtered_files.append(file_path)

    return filtered_files
