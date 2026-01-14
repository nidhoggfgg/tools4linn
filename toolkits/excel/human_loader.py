from __future__ import annotations

from logging import Logger, getLogger
from pathlib import Path
from typing import Optional, List, Dict

from openpyxl import load_workbook


class ExcelHumanLoader:
    """
    从 Excel 文件加载人员信息数据。
    支持从特定格式的 Excel 表格中提取日期、时间和人员信息。
    """

    def __init__(self, logger: Optional[Logger] = None):
        self.logger = logger or getLogger(__name__)

    def load_humans_from_excel(
        self,
        excel_file: Path | str,
        date_col: int = 0,  # A 列为日期，索引从 0 开始
        time_col: int = 1,  # B 列为开始结束时间
        name_col: int = 2,  # C 列为名称
        sheet_name: Optional[str] = None,
    ) -> List[Dict[str, str]]:
        """
        从 Excel 文件加载人员信息。

        Args:
            excel_file: Excel 文件路径
            date_col: 日期列索引（从 0 开始）
            time_col: 时间列索引（格式：xx:xx-xx:xx）
            name_col: 名称列索引
            sheet_name: 工作表名，None 表示激活的工作表

        Returns:
            人员信息列表，每个元素包含 date、start_time、end_time、human 字段

        格式说明：
        - 日期列格式为 x.xx，如 1.15 表示 01-15
        - 时间列格式为 xx:xx-xx:xx，如 09:00-17:00
        - 名称列可以包含多个名字，用空格、顿号、斜杠分隔，会统一转为 "and" 连接
        """
        input_path = Path(excel_file)
        if not input_path.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_path}")

        wb = load_workbook(str(input_path))
        ws = wb[sheet_name] if sheet_name else wb.active
        if ws is None:
            raise ValueError(f"工作表不存在")

        humans = []

        # 遍历所有行（跳过表头）
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # 跳过表头
                continue

            # 检查日期列是否为空，为空则停止
            if row[date_col] is None:
                break

            # 获取日期列的单元格对象，用于检查原始格式
            date_cell = ws.cell(row=i+1, column=date_col+1)

            try:
                # 解析日期：x.xx -> xx-xx
                # 注意：Excel 中的 12.10 会被读取为浮点数 12.1，需要根据数字格式判断
                date_value = row[date_col]

                if isinstance(date_value, str):
                    # 如果是字符串，直接替换
                    date = date_value.replace(".", "-")
                elif isinstance(date_value, (int, float)):
                    # 检查单元格的数字格式，判断原始是几位小数
                    number_format = date_cell.number_format

                    # 判断数字格式中小数点的位数
                    # 格式可能是 "0.00", "0.0", "General", "@", 等
                    decimal_places = 0
                    if "0.00" in number_format:
                        decimal_places = 2
                    elif "0.0" in number_format:
                        decimal_places = 1
                    elif "0" in number_format and "." in number_format:
                        # 计算小数点后有几位0
                        if "0.00" in number_format:
                            decimal_places = 2
                        elif "0.0" in number_format:
                            decimal_places = 1
                    else:
                        # 对于 General 格式，默认两位小数
                        decimal_places = 2

                    # 根据判断的小数位数格式化
                    if decimal_places == 2:
                        date = f"{date_value:.2f}".replace(".", "-")
                    elif decimal_places == 1:
                        int_part = int(date_value)
                        dec_part = round((date_value - int_part) * 10)
                        date = f"{int_part}-{dec_part:01d}"
                    else:
                        date = f"{date_value:.2f}".replace(".", "-")

                month, day = date.split("-")
                month = month.zfill(2)  # 补零
                day = day.zfill(2)
                date = f"{month}-{day}"

                # 解析时间：xx:xx-xx:xx
                time_str = str(row[time_col])
                # 处理可能的中文冒号
                time_str = time_str.replace("：", ":").replace(" ", "")
                start_time, end_time = time_str.split("-")
                start_time = start_time.strip().replace("：", ":")
                end_time = end_time.strip().replace("：", ":")

                # 解析名称：将分隔符统一为 "and"
                human = (
                    str(row[name_col])
                    .replace(" ", "and")
                    .replace("，", "and")
                    .replace("/", "and")
                )

                humans.append(
                    {
                        "date": date,
                        "start_time": start_time,
                        "end_time": end_time,
                        "human": human,
                    }
                )
            except Exception as e:
                self.logger.warning(f"解析第 {i+1} 行数据失败: {e}")
                continue

        self.logger.info(f"成功加载 {len(humans)} 条人员信息")
        return humans
