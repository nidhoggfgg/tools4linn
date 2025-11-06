from __future__ import annotations

import random
from datetime import datetime, timedelta
from logging import Logger, getLogger
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..time import generate_start_end_time
from ..utils import NameGenerator


class ExcelTimeFiller:
    """
    在保留 Excel 原有格式的前提下，填充开始时间、结束时间等列。
    支持同时填充名称列和持续时间列。
    支持随机调整行数，使数据看起来更加随机。
    """

    def __init__(self, logger: Optional[Logger] = None):
        self.logger = logger or getLogger(__name__)

    def fill_time_columns(
        self,
        excel_file: Path | str,
        sheet_name: Optional[str] = None,
        start_time_col: Optional[str | int] = None,  # 列字母，如 'A' 或列索引（从1开始）
        end_time_col: Optional[str | int] = None,  # 如 'B'
        duration_col: Optional[str | int] = None,  # 可选：持续时间列，公式为结束时间-开始时间
        start_time_range_start_str: str = "2025-01-01 00:00:00",  # 开始时间范围起始
        start_time_range_end_str: str = "2025-12-31 23:59:59",  # 开始时间范围结束
        end_time_range_start_str: str = "2025-01-01 00:00:00",  # 结束时间范围起始
        end_time_range_end_str: str = "2025-12-31 23:59:59",  # 结束时间范围结束
        data_start_row: int = 2,  # 数据从第几行开始（跳过标题）
        name_col: Optional[str | int] = None,  # 可选：名称列，如果提供则同时填充名称
        name_generator: Optional[NameGenerator] = None,  # 可选：名称生成器
        random_row_adjustment: tuple[int, int] = (0, 0),  # 随机行数调整范围
        output_file: Optional[Path | str] = None,
    ) -> Path:
        """
        在保留 Excel 原有格式的前提下，填充开始时间、结束时间两列。

        Args:
            excel_file: 输入的 Excel 文件路径（必须已存在，且有表头）
            sheet_name: 工作表名，None 表示激活的工作表
            start_time_col: 开始时间列标识，支持 'A' 或 1（从1开始）
            end_time_col: 结束时间列标识
            duration_col: 可选，持续时间列标识，设置公式为结束时间-开始时间
            start_time_range_start_str: 开始时间的随机范围起始
            start_time_range_end_str: 开始时间的随机范围结束
            end_time_range_start_str: 结束时间的随机范围起始
            end_time_range_end_str: 结束时间的随机范围结束
            data_start_row: 数据起始行（默认第2行，第1行为标题）
            name_col: 可选，名称列标识，如果提供则会在该列填充生成的名称
            name_generator: 可选，名称生成器实例，用于生成名称
            random_row_adjustment: 随机行数调整范围 (min, max)，负数表示删除行，正数表示增加行
            output_file: 输出文件路径，None 表示覆盖原文件

        Returns:
            输出文件路径
        """
        input_path = Path(excel_file)
        output_path = Path(output_file) if output_file else input_path

        if not input_path.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_path}")

        if start_time_col is None or end_time_col is None:
            raise ValueError("开始时间、结束时间列不能为空")

        # 解析时间范围
        start_time_range_start = datetime.fromisoformat(start_time_range_start_str)
        start_time_range_end = datetime.fromisoformat(start_time_range_end_str)
        end_time_range_start = datetime.fromisoformat(end_time_range_start_str)
        end_time_range_end = datetime.fromisoformat(end_time_range_end_str)

        if start_time_range_start >= start_time_range_end:
            raise ValueError("开始时间范围起始必须早于结束")
        if end_time_range_start >= end_time_range_end:
            raise ValueError("结束时间范围起始必须早于结束")

        # 加载工作簿
        wb = load_workbook(str(input_path))
        ws = wb[sheet_name] if sheet_name else wb.active
        if ws is None:
            raise ValueError(f"工作表 {sheet_name} 不存在")

        # 转换列标识为列字母
        st_col = self._to_col_letter(start_time_col)
        et_col = self._to_col_letter(end_time_col)
        dur_col = self._to_col_letter(duration_col) if duration_col is not None else None

        # 计算原始数据行数
        max_row = ws.max_row
        original_data_rows = max_row - data_start_row + 1

        # 随机调整行数
        adjusted_data_rows = self._adjust_row_count(
            ws, original_data_rows, random_row_adjustment, data_start_row
        )

        # 更新 max_row
        max_row = ws.max_row

        # 获取要跳过的列（时间相关列）
        skip_cols = {st_col, et_col}
        if dur_col:
            skip_cols.add(dur_col)

        # 复制模板行数据
        self._copy_template_data(ws, data_start_row, max_row, skip_cols)

        # 生成名称列表（如果需要）
        names = None
        name_col_letter = None
        if name_col is not None and name_generator is not None:
            name_col_letter = self._to_col_letter(name_col)
            if adjusted_data_rows > 0:
                names_needed = adjusted_data_rows + 10
                self.logger.info(
                    f"需要生成 {names_needed} 个名称（调整后数据行数：{adjusted_data_rows} + 10）"
                )
                names = name_generator.generate_names(names_needed)
                random.shuffle(names)
                names = names[:adjusted_data_rows]

        # 填充时间和名称数据
        self._fill_data(
            ws,
            data_start_row,
            max_row,
            st_col,
            et_col,
            dur_col,
            start_time_range_start,
            start_time_range_end,
            end_time_range_start,
            end_time_range_end,
            name_col_letter,
            names,
        )

        # 保存文件
        wb.save(str(output_path))

        filled_cols = []
        if dur_col is not None:
            filled_cols.append("持续时间")
        if name_col_letter is not None:
            filled_cols.append("名称")
        if filled_cols:
            self.logger.info(
                f"已填充时间和{'、'.join(filled_cols)}数据并保留格式，保存至: {output_path}"
            )
        else:
            self.logger.info(f"已填充时间数据并保留格式，保存至: {output_path}")

        return output_path

    def _to_col_letter(self, col: str | int) -> str:
        """将列标识统一转为列字母"""
        if isinstance(col, str):
            return col.upper()
        elif isinstance(col, int) and col > 0:
            return get_column_letter(col)
        else:
            raise ValueError(f"无效的列标识: {col}")

    def _adjust_row_count(
        self,
        ws,
        original_data_rows: int,
        random_row_adjustment: tuple[int, int],
        data_start_row: int,
    ) -> int:
        """调整工作表行数"""
        if random_row_adjustment[0] != 0 or random_row_adjustment[1] != 0:
            if random_row_adjustment[0] > random_row_adjustment[1]:
                raise ValueError("random_row_adjustment 的最小值不能大于最大值")
            adjustment = random.randint(random_row_adjustment[0], random_row_adjustment[1])
            adjusted_data_rows = max(0, original_data_rows + adjustment)
            self.logger.info(
                f"随机行数调整：原始{original_data_rows}行 -> 调整后{adjusted_data_rows}行 (调整值: {adjustment})"
            )
        else:
            adjusted_data_rows = original_data_rows
            adjustment = 0

        # 计算需要的总行数（包括标题行）
        target_total_rows = data_start_row + adjusted_data_rows - 1
        current_total_rows = ws.max_row

        if target_total_rows > current_total_rows:
            # 需要增加行
            ws[f"A{target_total_rows}"] = None
            rows_added = target_total_rows - current_total_rows
            self.logger.info(f"扩展了{rows_added}行，新max_row: {ws.max_row}")
        elif target_total_rows < current_total_rows:
            # 需要删除行
            rows_to_delete = current_total_rows - target_total_rows
            delete_start = target_total_rows + 1
            ws.delete_rows(delete_start, rows_to_delete)
            self.logger.info(f"删除了{rows_to_delete}行，新max_row: {ws.max_row}")

        return adjusted_data_rows

    def _copy_template_data(self, ws, data_start_row: int, max_row: int, skip_cols: set):
        """复制模板行数据到所有数据行"""
        template_row = data_start_row
        max_col = ws.max_column

        for row in range(data_start_row, max_row + 1):
            for col in range(1, max_col + 1):
                col_letter = get_column_letter(col)

                # 跳过时间相关列
                if col_letter in skip_cols:
                    continue

                source_cell = ws[f"{col_letter}{template_row}"]
                target_cell = ws[f"{col_letter}{row}"]

                # 复制值
                target_cell.value = source_cell.value

                # 复制样式
                try:
                    if hasattr(source_cell, "font") and source_cell.font:
                        target_cell.font = source_cell.font.copy()
                    if hasattr(source_cell, "border") and source_cell.border:
                        target_cell.border = source_cell.border.copy()
                    if hasattr(source_cell, "fill") and source_cell.fill:
                        target_cell.fill = source_cell.fill.copy()
                    if hasattr(source_cell, "number_format"):
                        target_cell.number_format = source_cell.number_format
                    if hasattr(source_cell, "alignment") and source_cell.alignment:
                        target_cell.alignment = source_cell.alignment.copy()
                except Exception:
                    pass  # 如果样式复制失败，跳过

        self.logger.info(f"已用第{template_row}行模板覆盖填充数据（跳过时间列）")

    def _fill_data(
        self,
        ws,
        data_start_row: int,
        max_row: int,
        st_col: str,
        et_col: str,
        dur_col: Optional[str],
        start_time_range_start: datetime,
        start_time_range_end: datetime,
        end_time_range_start: datetime,
        end_time_range_end: datetime,
        name_col_letter: Optional[str],
        names: Optional[list[str]],
    ):
        """填充时间和名称数据"""
        name_idx = 0
        for row in range(data_start_row, max_row + 1):
            # 生成开始时间和结束时间
            start_time, end_time = generate_start_end_time(
                start_time_range_start=start_time_range_start,
                start_time_range_end=start_time_range_end,
                end_time_range_start=end_time_range_start,
                end_time_range_end=end_time_range_end,
            )

            # 写入时间单元格
            start_cell = ws[f"{st_col}{row}"]
            end_cell = ws[f"{et_col}{row}"]

            start_cell.value = start_time
            end_cell.value = end_time

            # 复制时间列的对齐格式
            template_alignment = ws[f"{st_col}{data_start_row}"].alignment
            if template_alignment:
                start_cell.alignment = template_alignment.copy()
                end_cell.alignment = template_alignment.copy()

            # 如果提供了持续时间列，设置公式
            if dur_col is not None:
                duration_cell = ws[f"{dur_col}{row}"]
                duration_cell.value = f"={et_col}{row}-{st_col}{row}"

                # 复制持续时间列的格式
                template_duration_cell = ws[f"{dur_col}{data_start_row}"]
                try:
                    if hasattr(template_duration_cell, "number_format"):
                        duration_cell.number_format = template_duration_cell.number_format
                    if (
                        hasattr(template_duration_cell, "alignment")
                        and template_duration_cell.alignment
                    ):
                        duration_cell.alignment = template_duration_cell.alignment.copy()
                except Exception:
                    pass

            # 如果提供了名称列，填充名称
            if name_col_letter is not None and names is not None and name_idx < len(names):
                ws[f"{name_col_letter}{row}"].value = names[name_idx]
                name_idx += 1
